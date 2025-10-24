#!/usr/bin/env python3
"""
Simple Excel to CSV converter using openpyxl
"""

try:
    import openpyxl
    import csv
    import os
    
    def convert_excel_to_csv():
        """Convert Excel files to CSV using openpyxl"""
        
        excel_files = [
            'BOLT UBC First Byte - Stadium Operations.xlsx',
            'BOLT UBC First Byte - Merchandise Sales.xlsx', 
            'BOLT UBC First Byte - Fanbase Engagement.xlsx'
        ]
        
        for excel_file in excel_files:
            if os.path.exists(excel_file):
                print(f"Converting {excel_file}...")
                
                # Load workbook
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Create CSV filename
                csv_file = excel_file.replace('.xlsx', '.csv')
                
                # Write to CSV
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
                
                print(f"  → Saved as {csv_file}")
                
                # Get basic info
                print(f"  → Rows: {ws.max_row}")
                print(f"  → Columns: {ws.max_column}")
                print()
            else:
                print(f"File not found: {excel_file}")
    
    if __name__ == "__main__":
        convert_excel_to_csv()
        
except ImportError:
    print("openpyxl not available. Trying alternative method...")
    
    # Alternative: Use xlrd if available
    try:
        import xlrd
        import csv
        import os
        
        def convert_with_xlrd():
            excel_files = [
                'BOLT UBC First Byte - Stadium Operations.xlsx',
                'BOLT UBC First Byte - Merchandise Sales.xlsx', 
                'BOLT UBC First Byte - Fanbase Engagement.xlsx'
            ]
            
            for excel_file in excel_files:
                if os.path.exists(excel_file):
                    print(f"Converting {excel_file}...")
                    
                    # Note: xlrd doesn't support .xlsx files well, but let's try
                    try:
                        wb = xlrd.open_workbook(excel_file)
                        sheet = wb.sheet_by_index(0)
                        
                        csv_file = excel_file.replace('.xlsx', '.csv')
                        
                        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            for row_num in range(sheet.nrows):
                                writer.writerow(sheet.row_values(row_num))
                        
                        print(f"  → Saved as {csv_file}")
                        print(f"  → Rows: {sheet.nrows}")
                        print(f"  → Columns: {sheet.ncols}")
                        print()
                    except Exception as e:
                        print(f"  → Error: {e}")
                else:
                    print(f"File not found: {excel_file}")
        
        if __name__ == "__main__":
            convert_with_xlrd()
            
    except ImportError:
        print("Neither openpyxl nor xlrd available.")
        print("Please install one of them:")
        print("  pip install openpyxl")
        print("  or")
        print("  pip install xlrd")
